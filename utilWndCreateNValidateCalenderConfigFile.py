import tkinter as tk
from tkinter import filedialog as fd
from tkinter import messagebox
#from tkinter.tix import *

import pandas as pd
import openpyxl as excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Fill, Color


import csv
import sys
import os
from datetime import datetime
from io import StringIO
import re

cur_dir = ""
#sFileHeader = ""
ValidTimeframeInds = ['W','M','Q','S','A']
Valid_DOW_DOM = ['LW','LD','FW','FD','M-F','MON','TUE','WED','THU','FRI','SAT','SUN'] 
Valid_Months = ['JAN','FEB','MAR','APR','MAY','JUN','JUL','AUG','SEP','OCT','NOV','DEC'] 
Valid_Day_Codes = ['LW','LD','FW','FD'] 
Valid_Day_Occ = '^(SUN|MON|TUE|WED|THU|FRI|SAT)-(1|2|3|4|L|F)$'
Valid_FF_PRE_PROCESS = ['EXCEL','MF']

VALID_CAL_CONFIG_FILENAME="CalendarConfigFile.csv"
MAX_EXTRACT_ID_LEN=30
MAX_EXTRACT_DESC_LEN=40
MAX_FF_PRE_PROCESS_LEN=5
MAX_DELIVERY_LEN=20

"""
This program is used to create a .csv config file from the Calendar Config Excel file.
When creating the .csv file, this program will:

1) Remove the header row
2) Ensure there are no Pipe characters in the cell data
3) Ensure there are no comma characters in the cell data

"""

def sel_new_csv_file(lblLabel):

    global cur_dir

    filetypes = (
        ('CSV files', '*.csv'),
    )

    filename = fd.asksaveasfilename(
        title='SaveAs file',
        initialdir=cur_dir,
        filetypes=filetypes 
    )

    # Verify filename has extension
    # and that extension is csv
    arrFileParts = os.path.splitext(filename)
    if arrFileParts[1].strip() == "":
        filename = arrFileParts[0] + ".csv"
    elif arrFileParts[1].strip() != "csv":
        filename = arrFileParts[0] + ".csv"     

    # Verify that config filename is what system is expecting
    if os.path.basename(filename)  != VALID_CAL_CONFIG_FILENAME:
        messagebox.showerror("Error", f"Invalid Calendar Configuration filename. Filename must be {VALID_CAL_CONFIG_FILENAME}")
        return
        

    # Assign selected filename to screen label
    lblLabel.config(text=filename)

    # set the current default directory for dialog box
    cur_dir = os.path.dirname(filename)
    print (cur_dir)


def sel_xlsx_file(lblLabel):

    global cur_dir

    filetypes = (
        ('Excel files', '*.xlsx'),
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir=cur_dir,
        filetypes=filetypes 
    )


    # Verify filename has extension
    # and that extension is xlsx
    arrFileParts = os.path.splitext(filename)
    if arrFileParts[1].strip() == "":
        filename = arrFileParts[0] + ".xlsx"
    elif arrFileParts[1].strip() != "xlsx":
        filename = arrFileParts[0] + ".xlsx"     

    # Assign selected filename to screen label
    lblLabel.config(text=filename)

    # set the current default directory for dialog box
    cur_dir = os.path.dirname(filename)
    print (cur_dir)


def ValdiateCalendarConfigFileAction():

    global IN_EXCEL_CONFIG_FILE
    global OUT_DELIMITED_TXT_FILE

    ########################################
    # 1) Perform validation of input fields
    # 2) Perform Compare
    ########################################

    # load variables with window values
    txtInExcelFile = lblInFile1Text.cget("text")
    txtOutputCsvFile = lblOutFileText.cget("text")

    # Verify that input values have been entered/selected

    if txtInExcelFile == "":
        messagebox.showerror("Error", "Calendar config Excel file has not been selected!")
        return

    if txtOutputCsvFile == "":
        messagebox.showerror("Error", "CSV file to create has not been selected!")
        return

    #####################################################
    # 1) Load variables needed by compareFiles function
    #    using screen input values.
    # 2) Perform csv file comparison
    #####################################################
    IN_EXCEL_CONFIG_FILE = txtInExcelFile
    OUT_DELIMITED_TXT_FILE = txtOutputCsvFile
 
    validateExcelConfigFile()

    writeCSVFile()

    ###########################################
    # Comparison has completed
    ###########################################
    messagebox.showinfo("Complete", "Calendar config csv file has been created!")


def writeCSVFile():

    #########################################################
    # Create csv config file
    #########################################################
    InWrkbk = excel.load_workbook(IN_EXCEL_CONFIG_FILE)
    
    with open(OUT_DELIMITED_TXT_FILE, 'w', newline="") as f:  
        csvFile = csv.writer(f, delimiter='|')

        #########################################################
        # Process Input Excel file sheets 
        #########################################################
        for sheet in InWrkbk:

            print (f"sheetName: {sheet}")

            ###########################################    
            # Iterate thru rows
            ###########################################    
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

                ###############################################
                # Set skip header Row
                ###############################################
                if row[0].row == 1:
                    continue

                ###############################################
                # Process all cells in row to see if it is 
                #    and empty row
                ###############################################  
                bEmptyRow = True
                for cell in row:

                    # Keep track of empty cells to know if row is blank    
                    if cell.value != None:
                        bEmptyRow = False

                if bEmptyRow:
                    continue        

                ###############################################
                # write csv record using row cell values 
                ###############################################  
                # create csv file
                csvFile.writerow([cell.value for cell in row])
          


    return 0


def validateExcelConfigFile():

    #########################################################
    # Validate input Calendar Config Excel file 
    #########################################################
    InWrkbk = excel.load_workbook(IN_EXCEL_CONFIG_FILE)

    #########################################################
    # Process Input Excel file sheets 
    #########################################################
    for sheet in InWrkbk:

        print (f"sheetName: {sheet}")

        bSheetHasErrors=False

        redColorFill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
        whiteColorFill = PatternFill(start_color="FFFFFF", end_color="FF0000", fill_type = "solid")
        Valid_Day_Occ_re = re.compile(Valid_Day_Occ, re.IGNORECASE)

        ###########################################    
        # Iterate thru columns and cells
        # NOTE: Skip header record
        ###########################################    
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):

            # Set skip header Row
            if row[0].row == 1:
                continue

            ###############################################
            # Process all cells in row
            ###############################################  
            for cell in row:

                # initialize cell to while - no Errors
                cell.fill =  whiteColorFill 
 
                ###############################################
                # Timeframe entered and has valid value 
                ###############################################
                colLetter = cell.column_letter

                if colLetter == "C":
                    if str(cell.value) in ValidTimeframeInds:
                        pass
                    else:
                        # flag error with red color in cell 
                        bSheetHasErrors=True 
                        cell.fill =  redColorFill
                        continue 

                ###############################################
                # DOW_DOM entered and has valid value 
                ###############################################
                if colLetter == "D":
                    if str(cell.value).strip() == "" or cell.value == None:
                        continue   
                    else:    
                        # LW FW, etc.
                        if str(cell.value) in Valid_DOW_DOM:
                            continue

                        # FRI-2 ?
                        elif Valid_Day_Occ_re.match(str(cell.value).upper()):
                            continue

                        # is it a number?
                        elif str(cell.value).isdigit():
                            if int(cell.value) > 0 and int(cell.value) <= 31: 
                                continue
                            # end-if    
              
                        # If list, validate individual items
                        else:
                            # is it a list of values?
                            listItems = str(cell.value).strip().split(',')    
                            for listItem in listItems:
                                if listItem in Valid_DOW_DOM:     
                                   pass
                                else: 
                                    # flag error with red color in cell 
                                    bSheetHasErrors=True 
                                    cell.fill =  redColorFill
                                    # break out of for-loop
                                    break

                            # process next cell    
                            continue      

                ###############################################
                # Months entered and has valid value 
                ###############################################
                if colLetter == "E":
                    if str(cell.value).strip() == "" or cell.value == None:
                        continue   
                    else:    
                        if str(cell.value).upper() in Valid_Months:
                            continue
                        else:
                            # is it a list of values?
                            listItems = str(cell.value).strip().upper().split(',')    
                            for listItem in listItems:
                                if listItem in Valid_Months:     
                                   pass
                                else: 
                                    # flag error with red color in cell 
                                    bSheetHasErrors=True 
                                    cell.fill =  redColorFill
                                    # break out of for-loop
                                    break

                            # process next cell    
                            continue  


                ###############################################
                # Days entered and has valid value 
                ###############################################
                if colLetter == "F":
                    cell.fill = whiteColorFill

                    if str(cell.value).strip() == "" or cell.value == None:
                        continue   
                    # is it LW, FW, LD, FD ?
                    elif str(cell.value).upper() in Valid_Day_Codes:
                        continue
                    # is it a number?
                    elif str(cell.value).isdigit():
                        if int(cell.value) > 0 and int(cell.value) <= 31: 
                            continue
                        else:
                            # flag error with red color in cell 
                            bSheetHasErrors=True 
                            cell.fill =  redColorFill
                            continue  
        
                    else:      
                        # Is it ex. FRI-2 ?       
                        match = Valid_Day_Occ_re.match(str(cell.value).upper())
                        if match:
                            continue
                        else:    
                            # flag error with red color in cell 
                            bSheetHasErrors=True 
                            cell.fill =  redColorFill
                            continue  


                ###############################################
                # FF Pre-Process () 
                ###############################################
                if colLetter == "H":
                    if str(cell.value).strip() == "" or cell.value == None:
                        continue   
                    else:    
                        if len(str(cell.value)) > MAX_FF_PRE_PROCESS_LEN:
                            # flag error with red color in cell 
                            bSheetHasErrors=True 
                            cell.fill =  redColorFill
                            continue  


                ###############################################
                # Delivery Method () 
                ###############################################
                if colLetter == "I":
                    if str(cell.value).strip() == "" or cell.value == None:
                        continue   
                    else:    
                        if len(str(cell.value)) > MAX_DELIVERY_LEN:
                            # flag error with red color in cell 
                            bSheetHasErrors=True 
                            cell.fill =  redColorFill
                            continue 


                ###############################################
                # Verify that cell does not contain pipe char 
                ###############################################
                sCellValue = str(cell.value)
                if sCellValue.find('|') == -1:
                    pass
                else: 
                    # flag error with red color in cell 
                    bSheetHasErrors=True 
                    cell.fill =  redColorFill
                    continue

                ###############################################
                # Extract ID and Desc must be entered 
                ###############################################
                if colLetter == "A" :
                    if str(cell.value).strip() == "" or cell.value == None:
                        bSheetHasErrors=True
                        cell.fill =  redColorFill 
                    elif  len(str(cell.value).strip() ) > MAX_EXTRACT_ID_LEN:
                        bSheetHasErrors=True                        
                        cell.fill =  redColorFill                            


                if colLetter == "B" :
                    if str(cell.value).strip() == "" or cell.value == None:
                        cell.fill =  redColorFill 
                    elif  len(str(cell.value).strip() ) > MAX_EXTRACT_DESC_LEN:
                        bSheetHasErrors=True                        
                        cell.fill =  redColorFill  


                ###############################################
                # Two-cell dependent edits 
                ###############################################
                TimeframeValue = sheet.cell(row=row[0].row, column=3).value  
                DOW_DOMValue = sheet.cell(row=row[0].row, column=4).value 
                MonthsValue = sheet.cell(row=row[0].row, column=5).value
                DaysValue = sheet.cell(row=row[0].row, column=6).value

                # If Weekly/Monthly; there must be a value in the DOW_DOM cell
                if TimeframeValue == "W" or TimeframeValue == "M" :
                    if DOW_DOMValue == None or str(DOW_DOMValue).strip() == "": 
                        sheet.cell(row=row[0].row, column=4).fill =  redColorFill 

                # If Qtr/Semi-Annual/Annual; there must be a value in the Months and Days cell
                if TimeframeValue != "W" and TimeframeValue != "M" :
                    if MonthsValue == None or str(MonthsValue).strip() == "": 
                        sheet.cell(row=row[0].row, column=5).fill =  redColorFill 
                        bSheetHasErrors = True

                if TimeframeValue != "W" and TimeframeValue != "M" :
                    if DaysValue == None or str(DaysValue).strip() == "": 
                        sheet.cell(row=row[0].row, column=6).fill =  redColorFill 
                        bSheetHasErrors = True


    if bSheetHasErrors == True:

        InWrkbk.save(IN_EXCEL_CONFIG_FILE)

        sGeneralErrorMsg="""
"Calendar config Excel file has bad data cells.\n 
\n   1) Cells should not contain Pipes.         
\n   2) ExtractID and Description must be entered. 
\n   3) Max ExtractID length=30; Max ExtractDesc length=40  
\n   4) DOW_DOM and Months list should be delimited by commas.
\n   5) When Timeframe = 'W' or 'M' --> DOW_DOM must be entered.  
\n   6) When Timeframe = 'Q' or 'S' or 'A' --> Months and Days must be entered.    
\n   7) Max FF_PRE_PROCESS length=5  
\n   8) Max Delivery Method length=20  
\n\nPlease correct."
"""
        messagebox.showerror("Error", sGeneralErrorMsg )
        sys.exit(12)  

    return 0


def main():

    ####################################################
    # define variables
    ####################################################
    global MDIWnd
    global lblInFile1Text
    global lblInFile2Text
    global lblOutFileText
    global txtOmitCols
    global intChkbxIgnoreCase 

    ####################################################
    # Build window
    ####################################################
    MDIWnd= tk.Tk()
    MDIWnd.title("Validate Excel Calendar Config File")
    MDIWnd.geometry("1100x300")

    lblHdr = tk.Label(MDIWnd, text='Validate Excel Calendar Config File')
    lblHdr.config(font=('helvetica', 14))
    lblHdr.grid(row=0, column=3, columnspan=3, padx=5, pady=10)

    lblSpacer = tk.Label(MDIWnd, text="          ")
    lblSpacer.grid(row=0, column=0)

    ##############################
    # Select InFile1
    ##############################
    btnInFile1 = tk.Button(text='  Select File  ', command=lambda:sel_xlsx_file(lblInFile1Text), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnInFile1.grid(row=1, column=1, pady=6)

    #lblInFile1Label = tk.Label(MDIWnd, text='Input File 1:', bd=1, relief="sunken")
    lblInFile1Label = tk.Label(MDIWnd, text='Calendar Config Excel File:')
    lblInFile1Label.config(font=('helvetica', 10))
    lblInFile1Label.grid(row=1, column=2, sticky='e', pady=1)

    # Selected InFile1
    lblInFile1Text = tk.Label(MDIWnd, text="")
    lblInFile1Text.config(font=('helvetica', 10))
    lblInFile1Text.grid(row=1, column=3, sticky='w')

    ###############################
    # Output File Label
    ###############################
    btnOutFile = tk.Button(text='  Select File ', command=lambda:sel_new_csv_file(lblOutFileText), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnOutFile.grid(row=3, column=1, padx=2, pady=2)

    lblOutFileLabel = tk.Label(MDIWnd, text='Calendar Output CSV Config File for S3:')
    lblOutFileLabel.config(font=('helvetica', 10))
    lblOutFileLabel.grid(row=3, column=2, sticky='e')

    # Selected Output File
    lblOutFileText = tk.Label(MDIWnd, text='')
    lblOutFileText.config(font=('helvetica', 10))
    lblOutFileText.grid(row=3, column=3, sticky='w')

    ###############################
    # Buttons
    ###############################
    btnCreate = tk.Button(text='Validate Calendar Config file', command=ValdiateCalendarConfigFileAction, bg='blue', fg='white', font=('helvetica',10, 'bold'))
    btnCreate.grid(row=6, column=3, sticky='w', pady=10)

    ####################################
    # Process Window messages
    ####################################
    MDIWnd.mainloop()


if __name__ == "__main__":
    
    main()




