import tkinter as tk
from tkinter import filedialog as fd
from tkinter import messagebox
#from tkinter.tix import *

import pandas as pd
import openpyxl as excel
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Fill, Color

import os
from datetime import datetime
from io import StringIO

cur_dir = ""
#sFileHeader = ""

"""
PSA summary (prior): 

In summary, once we ran the extracts, we could download the .csv.gz files to our AWS workspace, unzip it as a .csv file, and load and save the file as an Excel file, but we had 2 “gotchas”.  The columns after the “key” column are in alphabetical order by description. However, we would need to 1) move a column for two of the spreadsheets as the second column (breaking the alphabetical order), and 2) add styles (adjust column widths, add header color) to match the prior year’s deliverable Excel file. This was easy to do in windows, but could not be accomplished in the AWS workspace using Calc (Excel substitute).  And, these were still manual steps.

PSA summary (now) to remove some of the manual effort:

1)	Modified two of the shell scripts to (with a hard-coded description for one column), to move the two columns mentioned above so that it won’t have to be done manually in Calc in the AWS workspace by the programmer. Note: if the description for the column-to-move changes, the column description won’t be found in the list of column descriptions, and won’t be moved. (So, this hinges on the description not changing).

2)	Created a Python program to copy the styles (limited functionality – just for this extract – could be expanded to be used for other extracts). The input file is the prior year’s extract, and the File-to-update is the current year’s extract file saved as an Excel file (still need to do this manual step).  After pressing the “Copy Styles” button, the Excel file to update should have the same styles as the original. (Tested this in AWS workspace successfully).

Also, including a procedure (word doc) outlining the steps to execute the PSA extract. With the delivery of the Excel files (zipped as a .gz file by Jag) to the client. I think this extract is done.


"""

def sel_xlsx_file(lblLabel):

    global cur_dir

    filetypes = (
        ('Excel files', '*.xlsx'),
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir=cur_dir,
        filetypes=filetypes 
    )


    # Verify filename has extension
    # and that extension is xlsx
    arrFileParts = os.path.splitext(filename)
    if arrFileParts[1].strip() == "":
        filename = arrFileParts[0] + ".xlsx"
    elif arrFileParts[1].strip() != "xlsx":
        filename = arrFileParts[0] + ".xlsx"     

    # Assign selected filename to screen label
    lblLabel.config(text=filename)

    # set the current default directory for dialog box
    cur_dir = os.path.dirname(filename)
    print (cur_dir)


def initiateUpdateStylesAction():

    global IN_EXCEL_MODEL_FILE
    global OUT_UPDATE_EXCEL_FILE

    ########################################
    # 1) Perform validation of input fields
    # 2) Perform Compare
    ########################################

    # load variables with window values
    txtInExcelFile = lblInFile1Text.cget("text")
    txtOutExcelFile = lblOutFileText.cget("text")

    # Verify that input values have been entered/selected

    if txtInExcelFile == "":
        messagebox.showerror("Error", "Input Excel Model file has not been selected!")
        return

    if txtOutExcelFile == "":
        messagebox.showerror("Error", "Excel file to update has not been selected!")
        return

    #####################################################
    # 1) Load variables needed by compareFiles function
    #    using screen input values.
    # 2) Perform csv file comparison
    #####################################################
    IN_EXCEL_MODEL_FILE = txtInExcelFile
    OUT_UPDATE_EXCEL_FILE = txtOutExcelFile
 
    updateExceFileStyles()

    ###########################################
    # Comparison has completed
    ###########################################
    messagebox.showinfo("Complete", "Excel file has been updated!")


def updateExceFileStyles():

    #########################################################
    # Load input and output Excel spreadsheets
    #########################################################
    InWrkbk = excel.load_workbook(IN_EXCEL_MODEL_FILE)
    OutWrkbk = excel.load_workbook(OUT_UPDATE_EXCEL_FILE)

    # Set Excel border value
    thin = Side(border_style="thin")
    cellBorder = Border(top=thin, left=thin, right=thin, bottom=thin)

   
    #########################################################
    # Process Input Excel file sheets 
    #########################################################
    for sheet in InWrkbk:

        print (f"sheetName: {sheet}")

        # Set Row Height of header row
        hdrRowHeight = sheet.row_dimensions[1].height
        print(f"hdrRowHeight:{hdrRowHeight}")

        outSheet = OutWrkbk.worksheets[0]
        outSheet.row_dimensions[1].height = hdrRowHeight

        ###########################################    
        # Iterate thru columns and cells
        ###########################################    
        for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column ):

            print("")
 
            for cell in col:

                ###############################################
                # InSheet - get first cell color in Header Row
                ###############################################
                if col[0].col_idx == 1:
                    #print(f" cell.fill.fgColor:{ cell.fill.fgColor.rgb}")
                    AlphaRGB = cell.fill.fgColor.rgb 
                    #print(f"type:{type(AlphaRGB)}")
                    RGB = AlphaRGB[2:]
                    print(f"RGB:{RGB}")
                    #RGB = '01B0F1'
                    cellColorFillHdr = PatternFill(start_color=RGB, end_color=RGB, fill_type = "solid") 

                ###############################################
                # Get InSheet column sizes 
                ###############################################
                colLetter = cell.column_letter
                newWidth = sheet.column_dimensions[cell.column_letter].width

                print(f"colLetter:{colLetter}")
                print(f"column width:{newWidth}")

                ###############################################
                # OutSheet - Set Color for each cell in header 
                ###############################################
                print(f"outSheet:{outSheet}")
                
                outCell = outSheet[colLetter+"1"]
                #outSheet.row_dimension[0].height = 50

                outCell.fill = cellColorFillHdr
                outCell.border = cellBorder

                outCell.alignment =  Alignment(wrap_text=True,vertical='top') 

                ###############################################
                # OutSheet - Set column sizes 
                ###############################################
                outSheet.column_dimensions[colLetter].width = newWidth  
                #outSheet.column_dimensions[colLetter].width = 8                 



        OutWrkbk.save(OUT_UPDATE_EXCEL_FILE)

        continue

    return 0


def main():

    ####################################################
    # define variables
    ####################################################
    global MDIWnd
    global lblInFile1Text
    global lblInFile2Text
    global lblOutFileText
    global txtOmitCols
    global intChkbxIgnoreCase 

    ####################################################
    # Build window
    ####################################################
    MDIWnd= tk.Tk()
    MDIWnd.title("Excel-2-Excel file Copy Styles")
    MDIWnd.geometry("1100x300")

    lblHdr = tk.Label(MDIWnd, text='Excel-2-Excel file Copy Styles')
    lblHdr.config(font=('helvetica', 14))
    lblHdr.grid(row=0, column=3, columnspan=3, padx=5, pady=10)

    lblSpacer = tk.Label(MDIWnd, text="          ")
    lblSpacer.grid(row=0, column=0)

    ##############################
    # Select InFile1
    ##############################
    btnInFile1 = tk.Button(text='  Select File  ', command=lambda:sel_xlsx_file(lblInFile1Text), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnInFile1.grid(row=1, column=1, pady=6)

    #lblInFile1Label = tk.Label(MDIWnd, text='Input File 1:', bd=1, relief="sunken")
    lblInFile1Label = tk.Label(MDIWnd, text='Excel Input Model File:')
    lblInFile1Label.config(font=('helvetica', 10))
    lblInFile1Label.grid(row=1, column=2, sticky='e', pady=1)

    # Selected InFile1
    lblInFile1Text = tk.Label(MDIWnd, text="")
    lblInFile1Text.config(font=('helvetica', 10))
    lblInFile1Text.grid(row=1, column=3, sticky='w')

    ###############################
    # Output File Label
    ###############################
    btnOutFile = tk.Button(text='  Select File ', command=lambda:sel_xlsx_file(lblOutFileText), bg='blue', fg='white', font=('helvetica', 8, 'bold'))
    btnOutFile.grid(row=3, column=1, padx=2, pady=2)

    lblOutFileLabel = tk.Label(MDIWnd, text='Excel File to Update:')
    lblOutFileLabel.config(font=('helvetica', 10))
    lblOutFileLabel.grid(row=3, column=2, sticky='e')

    # Selected Output File
    lblOutFileText = tk.Label(MDIWnd, text='')
    lblOutFileText.config(font=('helvetica', 10))
    lblOutFileText.grid(row=3, column=3, sticky='w')

    ###############################
    # Buttons
    ###############################
    btnCreate = tk.Button(text='Copy Styles', command=initiateUpdateStylesAction, bg='blue', fg='white', font=('helvetica',10, 'bold'))
    btnCreate.grid(row=6, column=3, sticky='w', pady=10)

    ####################################
    # Process Window messages
    ####################################
    MDIWnd.mainloop()


if __name__ == "__main__":
    
    main()




