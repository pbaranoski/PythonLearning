import os
import sys
import pandas
import openpyxl as excel
#
# from openpyxl.xml.constants import MAX_ROW
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill


in_dir = r"C:\Users\user\Documents\PythonLearning\Excel"
out_dir = r"C:\Users\user\Documents\PythonLearning\Excel"

in_file = os.path.join(in_dir,"Boguty parish index info.xlsx")
in_file2 = os.path.join(in_dir,"Boguty parish index info2.xlsx")
out_file = os.path.join(in_dir,"FamilyMember.csv")

###############################
# Set Excel styles
###############################
bold_font = Font(bold=True)
big_red_text = Font(color="00FF0000", size=20)
center_aligned_text = Alignment(horizontal="center")
double_border_side = Side(border_style="double")
square_border = Border(top=double_border_side, right=double_border_side,
                        bottom=double_border_side, left=double_border_side)

skyblueFill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')

#assert os.path.isfile(path)
#datestring = datetime.strftime(datetime.now(), ' %Y_%m_%d')

###################################################
# process Excel spreadsheet
###################################################
wrkbk = excel.load_workbook(in_file)

if wrkbk is None:
    print("couldn't get workbook onject")
else:
    print("Got the workbook")    

#nofWrkSheets = len(wrkbk.worksheets)
#for sheetName in wrkbk.sheetnames:
#    print(sheetName)

for sheet in wrkbk.worksheets:
    print("Sheet Name:"+ sheet.title)

# Get active sheet
#sheet = wrkbk.active

###################################################
# process Excel spreadsheet
###################################################
iMaxCols = 0
iMaxRows = 0

for sheet in wrkbk.worksheets:
    if sheet.title == "Jazwinski":
        print("Process Sheet")
        iMaxRows = sheet.max_row
    else:
        continue  

    # sheet max rows and cols -- includes many empty columns and rows    
    print("max row: "+str(sheet.max_row))
    #print("max col: "+str(sheet.max_column))

    ####################################
    # 1) get row 1 header column values
    # 2) get true NOF columns 
    ####################################
    for cell in sheet[1:1]:
        if not cell.value is None:
            print(cell.value)
            iMaxCols += 1

    print("True Max columns:" + str(iMaxCols))
    print (get_column_letter(iMaxCols))

    ###############################
    # Set Auto filter
    ###############################
    #sheet.auto_filter(0,0,iMaxRows, iMaxCols)
    ##sheet.filt .filter_column(2, 'Record Type == Birth')

    ###############################
    # iterate over rows 
    ###############################
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=iMaxCols):
        for cell in row:
            sCell = cell.value
            if sCell is None or str(sCell).strip() == "":
                pass
            else:
                print(sCell)                  

###############################
# Set styles
###############################
sheet = wrkbk["Jazwinski"]
print("current value: "+str(sheet["A1"].value))

sheet["A2"] = "Mickey Mouse"
print("current value: "+str(sheet["A2"].value))
sheet["A2"].fill = skyblueFill
sheet["A2"].font = bold_font
sheet["A2"].alignment = center_aligned_text
sheet["A2"].border = square_border

wrkbk.save(filename=in_file2)  # save the workbook
wrkbk.close()  # close the workbook

###################################################
# Create separate Excel spreadsheets 
###################################################  
dfExcel = pandas.read_excel(in_file, sheet_name="Jazwinski")
dfExcel.dropna(axis=0, subset=['Record Type'], inplace=True)

recTypes = dfExcel["Record Type"].unique()
  
for recType in recTypes:
    dfJaz = dfExcel[dfExcel["Record Type"].str.contains(recType)]
    dfJaz.to_excel(os.path.join(in_dir, "Jazwinski_"+recType+".xlsx"), index=False)      




